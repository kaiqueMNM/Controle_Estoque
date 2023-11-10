from tkinter import *
from tkinter import ttk
from tkinter import messagebox
import openpyxl
from openpyxl import Workbook
import pathlib

class EstoqueApp(Tk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.appearance()
        self.todo_sistema()

    def layout_config(self):
        self.title("Sistema de Controle de Estoque")
        self.geometry("900x860")

    def appearance(self):
        self.config(bg="#f0f0f0")

        self.lb_produto = Label(self, text="Produto", font=("Helvetica", 16), bg="#f0f0f0")
        self.lb_produto.place(x=50, y=320)

        self.produto_entry = Entry(self, width=30, font=("Helvetica", 16))
        self.produto_entry.place(x=50, y=350)

        self.lb_quantidade = Label(self, text="Quant", font=("Helvetica", 16), bg="#f0f0f0")
        self.lb_quantidade.place(x=300, y=320)

        self.quantidade_entry = Entry(self, width=10, font=("Helvetica", 16))
        self.quantidade_entry.place(x=300, y=350)

        self.lb_vpu = Label(self, text="VPU", font=("Helvetica", 16), bg="#f0f0f0")
        self.lb_vpu.place(x=450, y=320)

        # Ajuste para permitir vírgulas na entrada do VPU
        self.vpu_var = StringVar()
        self.vpu_entry = Entry(self, width=10, font=("Helvetica", 16), textvariable=self.vpu_var)
        self.vpu_entry.place(x=450, y=350)

        self.btn_adicionar = Button(self, text="Adicionar Produto", command=self.adicionar_produto, font=("Helvetica", 14), bg="#4CAF50", fg="white")
        self.btn_adicionar.place(x=50, y=400)

        self.resultado_text = Text(self, width=40, height=10, font=("Helvetica", 12), state="disabled")
        self.resultado_text.place(x=500, y=320)

        self.btn_consultar = Button(self, text="Consultar Estoque", command=self.consultar_estoque, font=("Helvetica", 14), bg="#008CBA", fg="white")
        self.btn_consultar.place(x=510, y=520)

        self.btn_limpar_consulta = Button(self, text="Limpar Consulta", command=self.limpar_consulta, font=("Helvetica", 14), bg="#FF5733", fg="white")
        self.btn_limpar_consulta.place(x=690, y=520)

    def todo_sistema(self):
        frame = Frame(self, width=900, height=50, bg="#333")
        frame.place(x=0, y=10)

        title = Label(frame, text="Sistema de Controle de Estoque", font=("Helvetica", 24), fg="white", bg="#333")
        title.place(x=250, y=10)

    def adicionar_produto(self):
        produto = self.produto_entry.get()
        quantidade = self.quantidade_entry.get()
        vpu = self.vpu_var.get()

        try:
            quantidade = int(quantidade)
            vpu = float(vpu.replace(",", "."))  # Substituir vírgulas por pontos para garantir a conversão correta
        except ValueError:
            messagebox.showerror("Erro", "A quantidade e o VPU devem ser números.")
            return

        ficheiro_path = pathlib.Path("Estoque.xlsx")
        if not ficheiro_path.exists():
            workbook = Workbook()
            folha = workbook.active
            folha['A1'] = "Produto"
            folha['B1'] = "Quant"
            folha['C1'] = "VPU"
            workbook.save(ficheiro_path)

        ficheiro = openpyxl.load_workbook('Estoque.xlsx')
        folha = ficheiro.active
        folha.append([produto, quantidade, vpu])
        ficheiro.save("Estoque.xlsx")

        messagebox.showinfo("Sistema", f"Produto '{produto}' adicionado ao estoque com sucesso!")

    def consultar_estoque(self):
        try:
            ficheiro = openpyxl.load_workbook('Estoque.xlsx')
            folha = ficheiro.active

            self.resultado_text.config(state="normal")
            self.resultado_text.delete(1.0, END)

            resultado = "\n".join([f"PRODUTO: {row[0]}\nQUANTIDADE: {row[1]}\nVPU: {row[2]}\n{'*' * 25}" for row in folha.iter_rows(values_only=True, min_row=2)])
            self.resultado_text.insert(END, resultado)
            
            self.resultado_text.config(state="disabled")

        except FileNotFoundError:
            messagebox.showerror("Erro", "O arquivo de estoque não foi encontrado. Adicione produtos primeiro.")

    def limpar_consulta(self):
        self.resultado_text.config(state="normal")
        self.resultado_text.delete(1.0, END)
        self.resultado_text.config(state="disabled")

    def change_apm(self, nova_aparencia):
        pass

if __name__ == "__main__":
    app = EstoqueApp()
    app.mainloop()
